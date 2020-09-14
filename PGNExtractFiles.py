import openpyxl
import os
import shutil

PATH='C:/Users/I823285/Desktop/Product documents/Cheetah/PBG Files extract.xlsx'
SEARCH_DIRECTORY = "C:/Users/I823285/Desktop/Product documents/Cheetah/ERP Project/Common"
TARGET_DIRECTORY = "C:/Users/I823285/Desktop/Product documents/Cheetah/ERP Project/Fixed Assets CommonFilesTest"
SHEET='FixedAssets'
li=[]
p=[]
invalid=[]
workbook=openpyxl.load_workbook(PATH)

sheet=workbook.get_sheet_by_name(SHEET)

rows=sheet.max_row
cols=sheet.max_column

for i in range(1,rows+1):
        if sheet.cell(row=i,column=1).value is None:
            continue
        else:
            x=sheet.cell(row=i, column=1).value
            y=x.strip()
            li.append(y)
print(li)
def find_all(filename, path):
    for root, dirs, files in os.walk(path):
        for name in files:
            if name == filename:
                return root+'/'+name

for j in range(len(li)):
    try:
        p.append(find_all(li[j], SEARCH_DIRECTORY))
        shutil.copy2(p[j], TARGET_DIRECTORY)

    except:
        print(li[j])
        #invalid.append ( (li[j])
        pass

print(len(p))
print(f"No Invalid files: {len(invalid)}")
print(invalid)
# Read a list of files provided through excel #DONE
# Search the files in Common folder path location
# Select these files one by one and move them to another folder